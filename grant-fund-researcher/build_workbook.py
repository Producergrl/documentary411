#!/usr/bin/env python3
"""Build the lead-tracking workbook for Grant and Fund Researcher.

Usage:
    python3 build_workbook.py --output "My Project Funders and Sponsors.xlsx" \
        --project-name "My Project" --scope both

Scopes:
    grants   -> Grant and Funder Leads, Niche Foundations
    sponsors -> Strategic Partners, Sponsors
    both     -> all four sheets
"""

import argparse
import sys
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit(
        "openpyxl is required. Install it with:  pip install openpyxl"
    )

GRANT_COLUMNS = [
    "Organization Name", "Type/Category", "Contact Person", "Title/Role",
    "Email Address", "Phone", "Website", "Funding Focus",
    "Typical Grant Range", "Outreach Status", "Priority", "Notes",
    "Date Added", "Source URL",
]

SPONSOR_COLUMNS = [
    "Organization Name", "Type/Category", "Contact Person", "Title/Role",
    "Email Address", "Phone", "Website", "Sponsor Angle",
    "Potential Value/Offer", "Outreach Status", "Priority", "Notes",
    "Date Added", "Source URL",
]

SHEETS_BY_SCOPE = {
    "grants": [
        ("Grant and Funder Leads", GRANT_COLUMNS),
        ("Niche Foundations", GRANT_COLUMNS),
    ],
    "sponsors": [
        ("Strategic Partners", SPONSOR_COLUMNS),
        ("Sponsors", SPONSOR_COLUMNS),
    ],
}
SHEETS_BY_SCOPE["both"] = SHEETS_BY_SCOPE["grants"] + SHEETS_BY_SCOPE["sponsors"]

HEADER_FILL = PatternFill("solid", fgColor="7C0A00")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def add_sheet(wb, title, columns):
    ws = wb.create_sheet(title=title)
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        width = max(14, min(len(name) + 6, 32))
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    return ws


def add_readme(wb, project_name, sheet_names):
    ws = wb.create_sheet(title="About", index=0)
    ws.column_dimensions["A"].width = 100
    lines = [
        f"{project_name} — Funders and Sponsors",
        f"Created {date.today().isoformat()} by Grant and Fund Researcher.",
        "",
        "This workbook is the single source of truth for outreach. The daily "
        "research task reads it before every run and never re-adds an "
        "organization that already appears here.",
        "",
        "Sheets: " + ", ".join(sheet_names),
        "",
        "You can edit Outreach Status, Priority, and Notes freely. "
        "Suggested Outreach Status values: New, Drafted, Sent, Replied, "
        "Declined, Funded, Excluded.",
    ]
    for row, text in enumerate(lines, start=1):
        cell = ws.cell(row=row, column=1, value=text)
        if row == 1:
            cell.font = Font(bold=True, size=14)
    return ws


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", required=True, help="Path for the .xlsx file")
    parser.add_argument("--project-name", required=True)
    parser.add_argument("--scope", required=True, choices=sorted(SHEETS_BY_SCOPE))
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    sheets = SHEETS_BY_SCOPE[args.scope]
    for title, columns in sheets:
        add_sheet(wb, title, columns)
    add_readme(wb, args.project_name, [t for t, _ in sheets])

    wb.save(args.output)
    print(f"Workbook saved: {args.output}")
    print("Sheets created: " + ", ".join(ws.title for ws in wb.worksheets))


if __name__ == "__main__":
    main()
