"""
Build and update the AIGrantAndFundResearcher .xlsx funder spreadsheet.
New rows are starred (★) and highlighted gold. All rows sorted A–Z by org name.
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_SHEET_NAME = "My Funder Contacts"

_COLUMNS = [
    "Organization Name",
    "Type / Category",
    "Contact Person",
    "Title / Role",
    "Email Address",
    "Phone",
    "Website",
    "Mission / Focus Area",
    "Grant Range",
    "Deadline / Cycle",
    "Priority",
    "Notes",
]

_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_NEW_ROW_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
_NEW_ROW_FONT_BOLD = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)

_COL_MIN_WIDTH = 15
_COL_MAX_WIDTH = 50


def append_and_save(
    funders: list,  # list[FunderResult]
    spreadsheet_path: str | Path,
) -> int:
    """
    Load (or create) the spreadsheet, append new funders, re-sort, save.
    Returns the number of rows appended.
    """
    path = Path(spreadsheet_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = openpyxl.load_workbook(path)
        if _SHEET_NAME in wb.sheetnames:
            ws = wb[_SHEET_NAME]
        else:
            ws = wb.active
            ws.title = _SHEET_NAME
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _SHEET_NAME
        _write_header(ws)

    # Collect all existing data rows (skip header row 1)
    existing_rows: list[list] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            existing_rows.append(list(row))

    # Build new rows from FunderResult objects
    new_rows: list[list] = []
    for f in funders:
        row = [""] * len(_COLUMNS)
        row[0] = f.org_name or ""
        row[1] = ""
        row[2] = f.contact_person or ""
        row[3] = f.title or ""
        row[4] = f.email or ""
        row[5] = ""
        row[6] = f.website or ""
        row[7] = (f.mission or "")[:500]  # cap long mission text
        row[8] = f.grant_range or ""
        row[9] = f.deadline or ""
        row[10] = ""
        row[11] = "★ New"
        new_rows.append(row)

    all_rows = existing_rows + new_rows
    # Sort A–Z by org name (column 0), case-insensitive
    all_rows.sort(key=lambda r: (r[0] or "").lower().lstrip("★ "))

    # Clear sheet below header and rewrite
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
            cell.font = _NORMAL_FONT

    ws.delete_rows(2, ws.max_row)

    new_org_names = {(r[0] or "").strip() for r in new_rows}

    for row_data in all_rows:
        ws.append(row_data)
        row_idx = ws.max_row
        is_new = (row_data[0] or "").strip() in new_org_names

        for col_idx in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            if is_new:
                cell.fill = _NEW_ROW_FILL
                if col_idx == 1:
                    cell.font = _NEW_ROW_FONT_BOLD
                else:
                    cell.font = _NORMAL_FONT
            else:
                cell.font = _NORMAL_FONT

    _apply_column_widths(ws)

    # Rewrite header formatting (rewrite after clearing may lose it)
    _write_header(ws, reformat_only=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(path)
    logger.info("Saved spreadsheet: %s (%d new rows)", path, len(new_rows))
    return len(new_rows)


def _write_header(ws, reformat_only: bool = False) -> None:
    if not reformat_only:
        ws.append(_COLUMNS)
    for col_idx, _ in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_column_widths(ws) -> None:
    for col_idx in range(1, len(_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(_COLUMNS[col_idx - 1])
        for row in ws.iter_rows(
            min_row=2, min_col=col_idx, max_col=col_idx, values_only=True
        ):
            for cell_val in row:
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), _COL_MAX_WIDTH))
        width = max(_COL_MIN_WIDTH, min(max_len + 2, _COL_MAX_WIDTH))
        ws.column_dimensions[col_letter].width = width
